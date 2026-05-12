from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from outputs.excel import ExcelOutput
from outputs.report_routes import resolve_report_output_path, report_sharepoint_url
from scripts.ops.aggregate_summaries import apply_payload, process_once


def _make_summary_workbook(base: Path) -> Path:
    overview_dir = base / 'Overzicht'
    overview_dir.mkdir(parents=True, exist_ok=True)
    wb_path = overview_dir / '[RSA] Overzicht rapporten.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Overzicht'
    ws['F3'] = 'rapportnummer'
    wb.save(wb_path)
    return wb_path


def test_process_once_keeps_rows_separate_per_report(tmp_path):
    output_dir = tmp_path / 'RSA_OneDrive'
    staged_dir = output_dir / 'staged_summaries'
    staged_dir.mkdir(parents=True, exist_ok=True)
    _make_summary_workbook(output_dir)

    payload_a = {
        'operation': 'write_cell',
        'excel_filename': '[RSA] Overzicht rapporten.xlsx',
        'sheet': 'Overzicht',
        'cell': 'C4',
        'value': ['2026-05-12 05:00:00', 10],
        'meta': {'report': 'report0221'},
    }
    payload_b = {
        'operation': 'write_cell',
        'excel_filename': '[RSA] Overzicht rapporten.xlsx',
        'sheet': 'Overzicht',
        'cell': 'C4',
        'value': ['2026-05-12 05:05:00', 20],
        'meta': {'report': 'report0222'},
    }

    (staged_dir / 'a.json').write_text(json.dumps(payload_a), encoding='utf-8')
    (staged_dir / 'b.json').write_text(json.dumps(payload_b), encoding='utf-8')

    processed = process_once(staged_dir, output_dir, limit=100, dry_run=False)

    assert processed == 2

    wb = load_workbook(output_dir / 'Overzicht' / '[RSA] Overzicht rapporten.xlsx')
    ws = wb['Overzicht']

    row_by_report = {}
    for row in range(4, ws.max_row + 1):
        report_name = ws.cell(row=row, column=6).value
        if report_name:
            row_by_report[str(report_name).strip().lower()] = row

    assert 'report0221' in row_by_report
    assert 'report0222' in row_by_report

    row_a = row_by_report['report0221']
    row_b = row_by_report['report0222']
    assert ws.cell(row=row_a, column=3).value == '2026-05-12 05:00:00'
    assert ws.cell(row=row_b, column=3).value == '2026-05-12 05:05:00'


def test_apply_payload_writes_hyperlink_without_formula(tmp_path):
    output_dir = tmp_path / 'RSA_OneDrive'
    _make_summary_workbook(output_dir)
    excel = ExcelOutput(output_dir=str(output_dir))

    payload = {
        'operation': 'write_cell',
        'excel_filename': '[RSA] Overzicht rapporten.xlsx',
        'sheet': 'Overzicht',
        'cell': 'B4',
        'value': {'display': 'Link', 'hyperlink': 'https://example.test/report.xlsx?web=1'},
        'meta': {'report': 'report0221'},
    }

    apply_payload(excel, payload, output_dir)

    wb = load_workbook(output_dir / 'Overzicht' / '[RSA] Overzicht rapporten.xlsx')
    ws = wb['Overzicht']

    row = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(row=r, column=6).value or '').strip().lower() == 'report0221':
            row = r
            break

    assert row is not None
    assert ws.cell(row=row, column=2).value == 'Link'
    assert ws.cell(row=row, column=2).hyperlink is not None
    assert ws.cell(row=row, column=2).hyperlink.target == 'https://example.test/report.xlsx?web=1'


def test_report_routes_append_xlsx_extension_when_missing(tmp_path):
    output_dir = tmp_path / 'RSA_OneDrive'
    routed = resolve_report_output_path(
        output_dir,
        excel_filename='[RSA] some report',
        report_name='report0228',
    )
    assert routed.name.endswith('.xlsx')

    link = report_sharepoint_url(
        excel_filename='[RSA] some report',
        report_name='report0228',
    )
    assert link is not None
    assert '.xlsx?web=1' in link


