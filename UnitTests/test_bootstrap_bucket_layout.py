from pathlib import Path
from openpyxl import Workbook, load_workbook
from scripts.ops.bootstrap_bucket_layout import ReportDescriptor, bootstrap_bucket_layout
def _make_report_workbook(path: Path, *, overzicht_rows: int, historiek_rows: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overzicht'
    ws.append(['header'])
    for idx in range(overzicht_rows):
        ws.append([f'row-{idx}'])
    hist = wb.create_sheet('Historiek')
    hist.append(['header'])
    for idx in range(historiek_rows):
        hist.append([f'hist-{idx}'])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
def _make_summary_workbook(path: Path, report_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overzicht'
    for _ in range(3):
        ws.append([''])
    ws.append(['', 'old link', '', '', '', report_name, '', ''])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
def test_bootstrap_moves_root_report_into_bucket_and_updates_summary(monkeypatch, tmp_path):
    output_dir = tmp_path / 'RSA_OneDrive'
    output_dir.mkdir()
    report = ReportDescriptor(
        report_name='report0002',
        report_title='TLCfipoorten hebben een sturingsrelatie naar een Verkeersregelaar',
        excel_filename='[RSA] TLCfipoorten hebben een sturingsrelatie naar een Verkeersregelaar.xlsx',
        report_number=2,
    )
    root_file = output_dir / report.excel_filename
    bucket_file = output_dir / '0000-0099' / report.excel_filename
    summary_file = output_dir / 'Overzicht' / '[RSA] Overzicht rapporten.xlsx'
    _make_report_workbook(root_file, overzicht_rows=8, historiek_rows=4)
    _make_report_workbook(bucket_file, overzicht_rows=1, historiek_rows=0)
    _make_summary_workbook(summary_file, report.report_name)
    monkeypatch.setattr('scripts.ops.bootstrap_bucket_layout.discover_report_descriptors', lambda: [report])
    stats, descriptors = bootstrap_bucket_layout(output_dir, dry_run=False)
    assert descriptors == [report]
    assert stats.discovered == 1
    assert stats.reports_replaced == 1
    assert stats.reports_migrated == 1
    assert not root_file.exists()
    assert bucket_file.exists()
    bucket_wb = load_workbook(bucket_file, read_only=True)
    try:
        assert 'Overzicht' in bucket_wb.sheetnames
        assert 'Historiek' in bucket_wb.sheetnames
        assert bucket_wb['Overzicht'].max_row >= 9
        assert bucket_wb['Historiek'].max_row >= 5
    finally:
        bucket_wb.close()
    summary_wb = load_workbook(summary_file, read_only=False)
    try:
        ws = summary_wb['Overzicht']
        cell = ws['B4']
        assert cell.value == 'Link'
        assert cell.hyperlink is not None
        assert '/0000-0099/' in str(cell.hyperlink.target)
    finally:
        summary_wb.close()
