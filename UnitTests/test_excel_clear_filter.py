from pathlib import Path
import sys

# ensure repo root on path
repo_root = Path(__file__).resolve().parents[1]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

from outputs.excel import ExcelOutput
import openpyxl


def test_clear_filter_removes_existing(tmp_path):
    out_dir = tmp_path / 'rsa_out'
    out_dir.mkdir()
    ex = ExcelOutput(output_dir=str(out_dir))

    wb_path = out_dir / 'test_clear_filter.xlsx'

    # create workbook with some data and set an autofilter manually
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resultaat'
    # metadata rows + header + data
    ws.append(['meta1'])
    ws.append(['meta2'])
    ws.append(['id', 'naam'])
    ws.append(['a1', 'r1'])
    ws.append(['a2', 'r2'])
    wb.save(wb_path)

    # manually set an auto_filter range and save
    wb2 = openpyxl.load_workbook(wb_path)
    ws2 = wb2['Resultaat']
    ws2.auto_filter.ref = 'A3:B5'
    wb2.save(wb_path)

    # ensure it's set
    wb_check = openpyxl.load_workbook(wb_path)
    assert wb_check['Resultaat'].auto_filter.ref == 'A3:B5'

    # call the clear_filter helper and verify it's removed
    ex.clear_filter(str(wb_path), 'Resultaat')

    wb_final = openpyxl.load_workbook(wb_path)
    # if attribute removed, getattr(..., 'ref', None) may raise; safer check:
    af = getattr(wb_final['Resultaat'], 'auto_filter', None)
    # either None or auto_filter with no ref
    assert af is None or getattr(af, 'ref', None) in (None, '')

