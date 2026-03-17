from pathlib import Path
import sys

# ensure repo root is on sys.path so imports like `outputs` resolve when tests are run
repo_root = Path(__file__).resolve().parents[1]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

from outputs.excel import ExcelOutput


def test_autofilter_is_set(tmp_path):
    out_dir = tmp_path / 'rsa_out'
    out_dir.mkdir()
    ex = ExcelOutput(output_dir=str(out_dir))

    wb_path = out_dir / 'test_autofilter.xlsx'
    # create a real workbook file so load_workbook can open it
    import openpyxl
    wb = openpyxl.Workbook()
    # ensure the default sheet is present; rename or create 'Resultaat'
    if 'Sheet' in wb.sheetnames:
        ws0 = wb['Sheet']
        ws0.title = 'Resultaat'
    else:
        wb.create_sheet('Resultaat')
    wb.save(wb_path)
    ex.create_sheet(wb_path, 'Resultaat', clear_if_exists=True)

    data = [
        ['meta1'],
        ['meta2'],
        ['uuid', 'naam'],
        ['a1', 'row1'],
        ['a2', 'row2'],
    ]

    # Use write_report to ensure high-level flow (it writes metadata, header, data)
    from datasources.base import QueryResult
    keys = ['id', 'naam']
    rows = [
        ['a1', 'row1'],
        ['a2', 'row2'],
    ]
    qr = QueryResult(keys=keys, rows=rows, last_data_update='2020-01-01')

    class Ctx:
        spreadsheet_id = str(wb_path)
        report_title = 'Test'
        datasource_name = 'ds'
        now_utc = '2020-01-01 00:00:00'
        excel_filename = wb_path.name

    ctx = Ctx()

    ex.write_report(ctx, qr, startcell='A1', add_filter=True)

    # verify auto_filter ref
    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Resultaat']
    assert ws.auto_filter is not None
    # header starts at row 3; two columns => A3:B5
    assert ws.auto_filter.ref == 'A3:B5'


