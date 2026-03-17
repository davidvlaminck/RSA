from pathlib import Path
import sys

# ensure repo root on path
repo_root = Path(__file__).resolve().parents[1]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

from outputs.excel import ExcelOutput
import openpyxl


def test_recalculate_formula_sets_calcprop(tmp_path):
    out_dir = tmp_path / 'rsa_out'
    out_dir.mkdir()
    ex = ExcelOutput(output_dir=str(out_dir))

    wb_path = out_dir / 'test_recalc.xlsx'

    # create workbook with a formula in Resultaat!A4
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resultaat'
    # meta rows and header
    ws.append(['meta1'])
    ws.append(['meta2'])
    ws.append(['val'])
    # put formula in A4
    ws['A4'] = '=1+1'
    wb.save(wb_path)

    # call recalculate via ExcelOutput helper
    ex.recalculate_formula(str(wb_path), 'Resultaat', 'A4')

    # reload and check workbook calc properties - try multiple attribute names
    wb2 = openpyxl.load_workbook(wb_path)
    calc_set = False
    try:
        calc_set = getattr(wb2.calc_properties, 'fullCalcOnLoad', False)
    except Exception:
        try:
            calc_set = getattr(wb2.calculation_properties, 'fullCalcOnLoad', False)
        except Exception:
            try:
                calc_set = getattr(wb2.properties, 'fullCalcOnLoad', False)
            except Exception:
                calc_set = False
    # openpyxl may not expose calcProps as attributes; instead inspect the archive directly
    import zipfile
    with zipfile.ZipFile(wb_path, 'r') as z:
        data = z.read('xl/workbook.xml')
    assert b'fullCalcOnLoad="1"' in data

