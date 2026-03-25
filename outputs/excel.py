from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Iterator, List
import datetime
import decimal
import shutil
import time
import importlib
import warnings
import logging
from contextlib import contextmanager

# lazy-loaded openpyxl handles
openpyxl = None
Workbook = None
WriteOnlyWorkbook = None
get_column_letter = None
load_workbook = None

def _ensure_openpyxl_loaded():
    global openpyxl, Workbook, WriteOnlyWorkbook, get_column_letter, load_workbook
    if openpyxl is not None:
        return
    try:
        # import and then apply narrow warning filters for known third-party deprecation messages
        openpyxl = importlib.import_module('openpyxl')
        # suppress specific DeprecationWarning emitted by openpyxl using datetime.utcnow()
        warnings.filterwarnings('ignore', message=r"datetime.datetime.utcnow\(\) is deprecated", category=DeprecationWarning)
        # suppress certifi/importlib-resources deprecation message seen in some environments
        warnings.filterwarnings('ignore', message=r"path is deprecated. Use files\(\) instead", category=DeprecationWarning)

        mod = openpyxl
        Workbook = getattr(mod, 'Workbook')
        load_workbook = getattr(mod, 'load_workbook')
        # WriteOnlyWorkbook location can vary; prefer writer.write_only
        # WriteOnlyWorkbook is in openpyxl.writer.write_only
        try:
            wmod = importlib.import_module('openpyxl.writer.write_only')
            WriteOnlyWorkbook = getattr(wmod, 'WriteOnlyWorkbook')
        except Exception:
            # fallback: try alternative path
            try:
                from openpyxl.writer.write_only import WriteOnlyWorkbook as _W
                WriteOnlyWorkbook = _W
            except Exception:
                WriteOnlyWorkbook = None
        try:
            utils_mod = importlib.import_module('openpyxl.utils')
            get_column_letter = getattr(utils_mod, 'get_column_letter')
        except Exception:
            try:
                from openpyxl.utils import get_column_letter as _g
                get_column_letter = _g
            except Exception:
                get_column_letter = None
    except Exception:
        openpyxl = None
        Workbook = None
        WriteOnlyWorkbook = None
        get_column_letter = None
        load_workbook = None
        raise


@contextmanager
def _file_lock(path: Path):
    """Context manager that acquires an advisory POSIX exclusive lock on a lockfile

    The lock file is created alongside the workbook (same directory) with suffix '.lock'.
    This uses fcntl.flock and blocks until the lock is acquired. Intended for POSIX only.
    """
    lock_path = Path(path).with_suffix(path.suffix + '.lock')
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    fh = None
    try:
        # open lock file for update/creation
        fh = open(lock_path, 'a+')
        try:
            import fcntl
            fcntl.flock(fh.fileno(), fcntl.LOCK_EX)
        except Exception:
            # if locking fails, close and raise
            fh.close()
            raise
        yield
    finally:
        try:
            if fh:
                try:
                    import fcntl
                    fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
                except Exception:
                    pass
                try:
                    fh.close()
                except Exception:
                    pass
        except Exception:
            pass


# module logger
logger = logging.getLogger(__name__)

from datasources.base import QueryResult
from .base import OutputWriteContext


class ExcelWriterError(Exception):
    pass


class ExcelOutput:
    """Write one .xlsx per report.

    The class offers a small helper API for creating workbooks/sheets and streaming rows.
    It is intentionally minimal for now (no formatting bells & whistles yet).
    """

    name = "Excel"

    def __init__(self, output_dir: str = "RSA_OneDrive"):
        # If a relative path is provided, resolve it against the repository root so
        # running scripts from inside `scripts/` doesn't create `scripts/RSA_OneDrive`.
        out = Path(output_dir)
        if not out.is_absolute():
            repo_root = Path(__file__).resolve().parents[1]
            out = repo_root / out
        self.output_dir = out.resolve()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # --- helpers -----------------------------------------------------------------
    def _normalize_value(self, val: Any) -> Any:
        if val is None:
            return ''
        if isinstance(val, decimal.Decimal):
            try:
                return float(val)
            except Exception:
                return str(val)
        if isinstance(val, datetime.datetime):
            # normalize datetimes to UTC and format as ISO (with 'T') without timezone
            # information to match expected summary format (YYYY-MM-DDTHH:MM:SS).
            try:
                if val.tzinfo is None:
                    # naive datetimes: assume UTC, preserve value but format with 'T'
                    return val.isoformat(timespec='seconds')
                else:
                    # aware datetimes: convert to UTC then drop tzinfo so string
                    # matches naive-ISO representation expected by callers/tests
                    dt_utc = val.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                    return dt_utc.isoformat(timespec='seconds')
            except Exception:
                return str(val)
        if isinstance(val, datetime.date):
            return val.strftime('%Y-%m-%d')
        return val

    def iter_rows_from_data(self, headers: List[str], rows: Iterable[Iterable[Any]]) -> Iterable[List[Any]]:
        """Yield normalized rows (including header row) lazily for an in-memory rows iterable.

        Backwards-compat helper: previous minimal implementation used this signature.
        """
        yield list(headers)
        for r in rows:
            yield [self._normalize_value(c) for c in r]

    def backup_file(self, workbook_path: Path) -> Path:
        # No-op: do not create backups by default. Backups are an explicit operational
        # action and can be created with scripts/create_backups.py when needed.
        # Keep function for API compatibility; return original path unchanged.
        return workbook_path

    def _atomic_save_workbook(self, wb, workbook_path: Path) -> None:
        """Save workbook to a temporary file and atomically replace the target path.

        This mirrors the write_single_cell pattern but for full workbooks.
        """
        import os
        from tempfile import NamedTemporaryFile

        dirpath = workbook_path.parent
        dirpath.mkdir(parents=True, exist_ok=True)
        with NamedTemporaryFile(prefix=workbook_path.stem + '_tmp_', suffix=workbook_path.suffix, dir=str(dirpath), delete=False) as tmp:
            tmp_name = tmp.name
        try:
            wb.save(tmp_name)
            # use a file lock while replacing the target to avoid races
            with _file_lock(workbook_path):
                os.replace(tmp_name, str(workbook_path))
            try:
                fd = os.open(str(workbook_path), os.O_RDONLY)
                try:
                    os.fsync(fd)
                finally:
                    os.close(fd)
            except Exception:
                pass
        finally:
            try:
                if os.path.exists(tmp_name):
                    os.remove(tmp_name)
            except Exception:
                pass

    def create_workbook_if_missing(self, workbook_path: Path) -> Path:
        workbook_path = Path(workbook_path)
        _ensure_openpyxl_loaded()
        # Create a minimal workbook file if it does not exist. Many write-paths call
        # this helper expecting a real .xlsx to be present before calling load_workbook.
        # Creation is performed atomically and is a best-effort operation.
        try:
            if workbook_path.exists():
                return workbook_path

            # ensure parent dir exists
            workbook_path.parent.mkdir(parents=True, exist_ok=True)

            # create a simple workbook with a default 'Overzicht' sheet and a
            # 'Historiek' sheet to mirror expected templates used by other helpers.
            try:
                wb = Workbook()
            except Exception:
                # fallback: ensure openpyxl loaded and re-try
                _ensure_openpyxl_loaded()
                wb = Workbook()

            # Rename default active sheet to 'Overzicht' if present
            try:
                if wb.sheetnames:
                    wb.active.title = 'Overzicht'
                else:
                    wb.create_sheet('Overzicht')
            except Exception:
                try:
                    wb.create_sheet('Overzicht')
                except Exception:
                    pass

            # ensure Historiek exists
            try:
                if 'Historiek' not in wb.sheetnames:
                    wb.create_sheet('Historiek')
            except Exception:
                pass

            # save atomically
            try:
                self._atomic_save_workbook(wb, workbook_path)
            except Exception:
                try:
                    wb.save(workbook_path)
                except Exception:
                    # if saving fails, propagate to caller
                    raise

            try:
                self._log_file_change(str(Path(workbook_path).resolve()), 'CREATED')
            except Exception:
                pass

            return workbook_path
        except Exception:
            # best-effort: return the path even if creation failed
            return workbook_path

    def create_sheet(self, workbook_path: Path, sheet_name: str, clear_if_exists: bool = True) -> None:
        workbook_path = Path(workbook_path)
        _ensure_openpyxl_loaded()
        # Do not implicitly create a workbook here; create_workbook_if_missing returns the path
        # but no longer creates files automatically (to avoid accidental overwrite).
        self.create_workbook_if_missing(workbook_path)
        if not workbook_path.exists():
            # nothing to do if workbook is missing
            return
        wb = load_workbook(workbook_path)
        if sheet_name in wb.sheetnames:
            # Avoid destructive clearing for important summary sheets.
            if clear_if_exists and sheet_name not in {"Overzicht", "Historiek"}:
                ws = wb[sheet_name]
                # remove all rows: easiest by recreating sheet
                idx = wb.sheetnames.index(sheet_name)
                wb.remove(ws)
                wb.create_sheet(sheet_name, idx)
            else:
                # keep existing content
                pass
        else:
            wb.create_sheet(sheet_name)
        wb.save(workbook_path)
        try:
            self._log_file_change(str(Path(workbook_path).resolve()), 'MODIFIED')
        except Exception:
            pass

    def delete_sheet(self, workbook_path: Path | str, sheet_name: str) -> None:
        """Delete a sheet from the workbook if it exists.

        Accepts a path or spreadsheet identifier (resolved via _resolve_workbook_path).
        This is a no-op when the workbook or sheet does not exist.
        """
        workbook_path = Path(workbook_path)
        # resolve if an identifier was provided relative to output_dir
        try:
            # if the provided path is not absolute and doesn't exist, attempt resolution
            if not workbook_path.exists():
                workbook_path = self._resolve_workbook_path(str(workbook_path))
        except Exception:
            # fallback: treat as Path
            workbook_path = Path(workbook_path)

        if not workbook_path.exists():
            return

        _ensure_openpyxl_loaded()
        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            return
        # remove the sheet
        ws = wb[sheet_name]
        wb.remove(ws)

        # atomic save
        try:
            self._atomic_save_workbook(wb, workbook_path)
        except Exception:
            try:
                wb.save(workbook_path)
            except Exception:
                raise

        try:
            self._log_file_change(str(Path(workbook_path).resolve()), 'MODIFIED')
        except Exception:
            pass

    def iter_rows(self, a, b=None) -> Iterator[list[Any]]:
        """Dual-purpose iter_rows:

        - Modern signature: iter_rows(workbook_path: Path, sheet_name: str) -> Iterator[list]
        - Legacy/testing signature: iter_rows(headers: List[str], rows: Iterable[Iterable[Any]]) -> Iterator[list]

        We detect by type of `a`.
        """
        # Legacy use: headers + rows
        if isinstance(a, (list, tuple)):
            headers = list(a)
            rows = b
            yield headers
            for r in rows:
                yield [self._normalize_value(c) for c in r]
            return

        # Otherwise treat as (workbook_path, sheet_name)
        workbook_path = Path(a)
        sheet_name = b
        if not workbook_path.exists():
            return
            yield  # pragma: no cover
        _ensure_openpyxl_loaded()
        wb = load_workbook(workbook_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            return
            yield  # pragma: no cover
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            yield [self._normalize_value(c) for c in row]

    def clear_sheet_range(self, workbook_path: Path, sheet_name: str, cell_range: str) -> None:
        workbook_path = Path(workbook_path)
        if not workbook_path.exists():
            return
        _ensure_openpyxl_loaded()
        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        # naive clear: iterate cells in range
        try:
            cells = ws[cell_range]
            # cells can be tuple of tuples
            for r in cells:
                for c in r:
                    c.value = None
            wb.save(workbook_path)
        except Exception:
            # fallback: recreate sheet
            ws = wb[sheet_name]
            idx = wb.sheetnames.index(sheet_name)
            wb.remove(ws)
            wb.create_sheet(sheet_name, idx)
            wb.save(workbook_path)
        try:
            self._log_file_change(str(Path(workbook_path).resolve()), 'MODIFIED')
        except Exception:
            pass

    # --- compatibility helpers that mimic SheetsWrapper (read-focused) -----------------
    def _resolve_workbook_path(self, spreadsheet_id_or_path: str | Path) -> Path:
        """Resolve a spreadsheet identifier to a workbook Path inside the configured output dir.

        Logic:
        - If the provided value is an absolute or relative path to an existing file, return it.
        - Else, look inside self.output_dir for the filename as-is or with .xlsx appended.
        - Fall back to treating the identifier as a filename under output_dir.
        """
        sp = str(spreadsheet_id_or_path)
        p = Path(sp)
        # If an absolute path was provided, treat it as the intended target even
        # if the file does not yet exist. This avoids accidentally constructing a
        # candidate under self.output_dir with the absolute path string embedded.
        if p.is_absolute():
            return p
        if p.exists():
            return p
        # try as file in output dir
        candidate = Path(self.output_dir) / sp
        if candidate.exists():
            return candidate
        # If the caller supplied what looks like a filename (has .xlsx suffix or any suffix),
        # prefer to return the candidate path (which may not exist yet). This avoids the
        # discovery scan picking a different workbook in output_dir that happens to
        # contain an Overzicht/Historiek sheet (which previously caused writes to end
        # up in another file). Returning the candidate allows writes to create the
        # intended workbook under output_dir.
        if p.suffix:
            return candidate
        # try mapping spreadsheet id -> filename (log any issues)
        try:
            from outputs.spreadsheet_map import lookup
            mapped = lookup(sp)
            logger.debug('Spreadsheet id lookup for "%s" -> %s', sp, mapped)
            if mapped:
                candidate_map = Path(self.output_dir) / mapped
                logger.debug('Candidate mapped path: %s (exists=%s)', candidate_map, candidate_map.exists())
                if candidate_map.exists():
                    return candidate_map
        except Exception as ex:
            logger.exception('Failed to lookup spreadsheet mapping for %s: %s', sp, ex)
        # If input had no suffix, try candidate with .xlsx appended.
        candidate_x = Path(self.output_dir) / (sp + '.xlsx') if not p.suffix else candidate

        # If no exact file was found, try to discover an existing summary workbook in output_dir.
        # This prevents creating a new file named after the spreadsheet id which would hide/replace
        # the original '[RSA] Overzicht rapporten.xlsx' workbook.
        try:
            if Path(self.output_dir).exists():
                for p in sorted(Path(self.output_dir).glob('*.xlsx')):
                    try:
                        _ensure_openpyxl_loaded()
                        wb_tmp = load_workbook(p, read_only=True)
                        if 'Overzicht' in wb_tmp.sheetnames or 'Historiek' in wb_tmp.sheetnames:
                            # register mapping for future fast lookup if possible (non-persistent)
                            try:
                                from outputs.spreadsheet_map import add_mapping
                                add_mapping(sp, p.name, persist=False)
                            except Exception:
                                pass
                            return p
                    except Exception:
                        continue
        except Exception:
            pass

        return candidate_x

    def get_sheets_in_spreadsheet(self, spreadsheet_id: str) -> dict:
        """Return a dict of sheet names -> properties similar to Sheets API minimal shape.

        Example return format:
        {
            'Sheet1': {'gridProperties': {'rowCount': 123, 'columnCount': 10}},
            ...
        }
        """
        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return {}
        wb = load_workbook(wb_path, read_only=True)
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            # best-effort grid properties
            try:
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0
            except Exception:
                row_count = 0
                col_count = 0
            out[name] = {'gridProperties': {'rowCount': row_count, 'columnCount': col_count}}
        return out

    def read_data_from_sheet(self, spreadsheet_id: str, sheet_name: str, sheetrange: str | None = None, *,
                             return_raw_results: bool = False, value_render_option: str ='FORMATTED_VALUE') -> list[list[Any]] | dict:
        """Return a list of rows (lists) for the requested range. If sheetrange is None, returns all rows.

        Accepts ranges like 'A1:C10', 'A1:' (until the last row), 'A' or 'A1' (single column), or full None.
        If `return_raw_results` is True, return a dict similar to Sheets API: {'values': rows, 'range': 'Sheet!A1:C3'}
        """
        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return {} if return_raw_results else []
        wb = load_workbook(wb_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            return {} if return_raw_results else []
        ws = wb[sheet_name]

        from openpyxl.utils import range_boundaries

        # Determine boundaries
        if sheetrange is None or sheetrange == '':
            min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
        else:
            # handle patterns like 'A1:' or 'A:' -> treat as until max
            if ':' in sheetrange:
                left, right = sheetrange.split(':', 1)
                if left == '' and right == '':
                    min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
                else:
                    if right == '' or right is None:
                        # open range, compute right using ws max
                        try:
                            min_col, min_row, _, _ = range_boundaries(left + ':' + left)
                        except Exception:
                            # fallback to column letter
                            from openpyxl.utils import column_index_from_string
                            col = ''.join([c for c in left if c.isalpha()])
                            min_col = column_index_from_string(col) if col else 1
                            min_row = int(''.join([c for c in left if c.isdigit()]) or 1)
                        max_col, max_row = ws.max_column or min_col, ws.max_row or min_row
                    else:
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(sheetrange)
                        except Exception:
                            # if parsing fails, fallback to whole sheet
                            min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
            else:
                # single column like 'A' or single cell 'A1'
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(sheetrange + ':' + sheetrange)
                except Exception:
                    min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1

        rows_out = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            rows_out.append([self._normalize_value(c) for c in row])

        if return_raw_results:
            # build a simple range string
            rng = sheetrange or (f'A1:{chr(64 + (len(rows_out[0]) if rows_out else 1))}{len(rows_out) if rows_out else 1}')
            return {'values': rows_out, 'range': f'{sheet_name}!{rng}'}
        return rows_out

    def read_celldata_from_sheet(self, spreadsheet_id: str, sheet_name: str, sheetrange: str | None = None, return_raw_results: bool = True) -> dict:
        """Return a minimal raw-style dict with 'values' and 'range' to mimic Sheets API basic shape.

        For now we return only values; rich `rowData` with hyperlinks/formulas can be added later.
        """
        rows = self.read_data_from_sheet(spreadsheet_id, sheet_name, sheetrange)
        # compute a simple range string
        rng = sheetrange or (f'A1:{chr(64 + (len(rows[0]) if rows else 1))}{len(rows) if rows else 1}')
        result: dict = {'values': rows, 'range': f'{sheet_name}!{rng}'}

        # build a richer rowData structure similar to the Sheets API so legacy code can access hyperlinks and formatted values
        try:
            wb_path = self._resolve_workbook_path(spreadsheet_id)
            if wb_path.exists():
                wb = load_workbook(wb_path, read_only=True)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rowData = []
                    # determine boundaries for iteration used earlier
                    from openpyxl.utils import range_boundaries
                    if sheetrange is None or sheetrange == '':
                        min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
                    else:
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(sheetrange)
                        except Exception:
                            min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1

                    for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                        values = []
                        for c in r:
                            cell_info: dict = {}
                            # userEnteredValue: raw python value
                            cell_info['userEnteredValue'] = c.value
                            # formattedValue: string representation
                            try:
                                cell_info['formattedValue'] = str(c.value) if c.value is not None else ''
                            except Exception:
                                cell_info['formattedValue'] = ''
                            # hyperlink if present
                            try:
                                if hasattr(c, 'hyperlink') and c.hyperlink:
                                    cell_info['hyperlink'] = str(c.hyperlink.target if hasattr(c.hyperlink, 'target') else c.hyperlink)
                            except Exception:
                                pass
                            # formula detection: in openpyxl formula is string starting with '=' or c.data_type
                            try:
                                if isinstance(c.value, str) and c.value.startswith('='):
                                    cell_info['formula'] = c.value
                            except Exception:
                                pass
                            values.append(cell_info)
                        rowData.append({'values': values})
                    result['rowData'] = rowData
        except Exception:
            # best-effort: if rich read fails, return only simple 'values'
            pass

        if return_raw_results:
            return result
        return result['values']

    def write_single_cell(self, spreadsheet_id: str | Path, sheet_name: str, cell: str, value: Any) -> None:
        """Write a single cell efficiently and atomically.

        - Accepts spreadsheet identifier or path.
        - Performs atomic save (write to temp file + os.replace) to avoid partial writes.
        """
        import os
        from tempfile import NamedTemporaryFile

        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        workbook_existed = wb_path.exists()
         # ensure workbook exists
        if not wb_path.exists():
            self.create_workbook_if_missing(wb_path)

        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # simple validation of cell
        col_letters = ''.join([c for c in cell if c.isalpha()])
        row_digits = ''.join([c for c in cell if c.isdigit()])
        if not col_letters or not row_digits:
            raise ExcelWriterError(f'Invalid cell identifier: {cell}')

        # normalize value consistently with other write paths
        try:
            norm_val = self._normalize_value(value)
        except Exception:
            norm_val = value
        ws[cell].value = norm_val

        # atomic save: write to temp file and replace under lock
        dirpath = wb_path.parent
        with NamedTemporaryFile(prefix=wb_path.stem + '_tmp_', suffix=wb_path.suffix, dir=str(dirpath), delete=False) as tmp:
            tmp_name = tmp.name
        try:
            wb.save(tmp_name)
            with _file_lock(wb_path):
                os.replace(tmp_name, str(wb_path))
            try:
                fd = os.open(str(wb_path), os.O_RDONLY)
                try:
                    os.fsync(fd)
                finally:
                    os.close(fd)
            except Exception:
                pass
        finally:
            try:
                if os.path.exists(tmp_name):
                    os.remove(tmp_name)
            except Exception:
                pass
        # log change
        try:
            action = 'MODIFIED' if workbook_existed else 'CREATED'
            self._log_file_change(str(Path(wb_path).resolve()), action)
        except Exception:
            pass

    def update_row_by_adding_number(self, spreadsheet_id: str | Path, sheet_name: str, cell: str, delta: int) -> None:
        """Read integer from `cell`, add `delta`, and write it back using atomic save.

        This method performs read-modify-write under the same lock policy as write_single_cell: if
        locking is enabled for this workbook, it will acquire the inter-process lock for the whole
        operation. Otherwise it will do a best-effort read+write with atomic replace.
        """
        import os
        from tempfile import NamedTemporaryFile

        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        workbook_existed = wb_path.exists()

         # ensure workbook exists
        if not wb_path.exists():
            self.create_workbook_if_missing(wb_path)

        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        try:
            val = ws[cell].value
            current = int(val) if (val is not None and str(val).strip() != '') else 0
        except Exception:
            current = 0

        new_val = current + int(delta)
        ws[cell].value = new_val

        # atomic save under file lock to prevent concurrent writers
        dirpath = wb_path.parent
        with NamedTemporaryFile(prefix=wb_path.stem + '_tmp_', suffix=wb_path.suffix, dir=str(dirpath), delete=False) as tmp:
            tmp_name = tmp.name
        try:
            wb.save(tmp_name)
            with _file_lock(wb_path):
                os.replace(tmp_name, str(wb_path))
            try:
                fd = os.open(str(wb_path), os.O_RDONLY)
                try:
                    os.fsync(fd)
                finally:
                    os.close(fd)
            except Exception:
                pass
        finally:
            try:
                if os.path.exists(tmp_name):
                    os.remove(tmp_name)
            except Exception:
                pass
        # log change
        try:
            action = 'MODIFIED' if workbook_existed else 'CREATED'
            self._log_file_change(str(Path(wb_path).resolve()), action)
        except Exception:
            pass

    def find_first_nonempty_row_from_starting_cell(self, spreadsheet_id: str | Path, sheet_name: str, start_cell: str, max_rows: int = 1000000) -> int:
        """Find the first non-empty row in the column of start_cell, searching up to max_rows rows.

        Returns the row number (1-based). If none found, returns ws.max_row + 1 or start_row if sheet missing.
        """
        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return 1
        wb = load_workbook(wb_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            return 1
        ws = wb[sheet_name]

        # parse start_cell like 'A1'
        col_letters = ''.join([c for c in start_cell if c.isalpha()])
        row_digits = ''.join([c for c in start_cell if c.isdigit()])
        if not col_letters:
            col_index = 1
        else:
            from openpyxl.utils import column_index_from_string
            col_index = column_index_from_string(col_letters)
        start_row = int(row_digits) if row_digits else 1

        max_row = ws.max_row or start_row
        limit = min(max_row, start_row + max_rows - 1)
        for r in range(start_row, limit + 1):
            try:
                val = ws.cell(row=r, column=col_index).value
            except Exception:
                val = None
            if val is not None and val != '':
                return r
        # if none found, return next row after existing max
        return max_row + 1

    def insert_empty_rows(self, spreadsheet_id: str | Path, sheet_name: str, start_cell: str = 'A2', number_of_rows: int = 1) -> None:
        """Insert empty rows at start_cell by shifting existing rows down.

        This is a simple implementation: read the sheet, insert rows in-memory, and save.
        """
        _ensure_openpyxl_loaded()
        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            # nothing to do
            return
        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            # create if missing
            wb.create_sheet(sheet_name)
            wb.save(wb_path)
            return
        ws = wb[sheet_name]
        # parse start row
        col_letters = ''.join([c for c in start_cell if c.isalpha()])
        row_digits = ''.join([c for c in start_cell if c.isdigit()])
        start_row = int(row_digits) if row_digits else 1
        ws.insert_rows(start_row, amount=number_of_rows)
        # atomic save
        try:
            self._atomic_save_workbook(wb, wb_path)
        except Exception:
            wb.save(wb_path)
        # log modification
        try:
            self._log_file_change(str(Path(wb_path).resolve()), 'MODIFIED')
        except Exception:
            pass

    # --- core write function ----------------------------------------------------
    def write_data_to_sheet(self, workbook_path: Path, sheet_name: str, rows: Iterable[Iterable[Any]],
                            start_cell: str = 'A1', overwrite: bool = True, use_write_only_for_nrows: int = 10000) -> dict:
        """Write rows to a sheet. Rows can be generator. Returns metadata dict."""
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            raise ExcelWriterError('openpyxl not available')

        workbook_path = Path(workbook_path)
        # ensure parent dir
        workbook_path.parent.mkdir(parents=True, exist_ok=True)

        # Note: do not create a backup file before write. We perform an atomic save
        # (write to temp file + os.replace) so a partial write does not corrupt the
        # existing workbook. If you still want backups, run the cleanup/backup script separately.

        # Try to write using WriteOnly if rows is a generator and large
        # We don't know rowcount ahead of time; detect generators by lack of __len__
        is_generator = not hasattr(rows, '__len__')

        start_time = time.time()
        rows_written = 0

        workbook_exists = workbook_path.exists()

        if workbook_exists:
            # Load existing workbook and replace/clear the target sheet to preserve other sheets
            _ensure_openpyxl_loaded()
            wb = load_workbook(workbook_path)
            if sheet_name in wb.sheetnames:
                # remove existing sheet and recreate at same position
                idx = wb.sheetnames.index(sheet_name)
                ws = wb[sheet_name]
                wb.remove(ws)
                ws = wb.create_sheet(sheet_name, idx)
            else:
                ws = wb.create_sheet(sheet_name)

            # write rows into the ws (works for generator and sequences)
            for r in rows:
                row = [self._normalize_value(v) for v in r]
                ws.append(row)
                rows_written += 1

            # atomic save under file lock
            try:
                # _atomic_save_workbook already acquires lock around replace
                self._atomic_save_workbook(wb, workbook_path)
            except Exception:
                try:
                    with _file_lock(workbook_path):
                        wb.save(workbook_path)
                except Exception:
                    raise

        else:
            # workbook does not exist yet: create a workbook with default template sheets
            # so that Overzicht and Historiek are present (preserve expected templates).
            # Prefer creating the template workbook first, then load and write the Resultaat sheet.
            self.create_workbook_if_missing(workbook_path)

            # now load the workbook and write the Resultaat sheet (works for generators and sequences)
            _ensure_openpyxl_loaded()
            wb = load_workbook(workbook_path)
            if sheet_name in wb.sheetnames:
                idx = wb.sheetnames.index(sheet_name)
                ws = wb[sheet_name]
                wb.remove(ws)
                ws = wb.create_sheet(sheet_name, idx)
            else:
                ws = wb.create_sheet(sheet_name)

            for r in rows:
                row = [self._normalize_value(v) for v in r]
                ws.append(row)
                rows_written += 1

            # atomic save under file lock
            try:
                self._atomic_save_workbook(wb, workbook_path)
            except Exception:
                try:
                    with _file_lock(workbook_path):
                        wb.save(workbook_path)
                except Exception:
                    raise

        elapsed = time.time() - start_time
        # log the file change: CREATED if it did not exist at start, else MODIFIED
        try:
            action = 'MODIFIED' if workbook_exists else 'CREATED'
            self._log_file_change(str(Path(workbook_path).resolve()), action)
        except Exception:
            pass
        return {'rows_written': rows_written, 'elapsed_seconds': elapsed, 'file': str(workbook_path)}

    # --- compatibility entrypoint used by DQReport --------------------------------
    def write_report(self, ctx: OutputWriteContext, result: QueryResult, *,
                     startcell: str = "A1",
                     add_filter: bool = True,
                     persistent_column: str = "",
                     persistent_dict: dict[str, str] | None = None,
                     convert_columns_to_numbers: list[str] | None = None,
                     link_type: str = "awvinfra",
                     recalculate_cells: list[tuple[str, str]] | None = None) -> None:
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            raise ExcelWriterError('openpyxl required')

        # determine output filename: prefer ctx.excel_filename if provided
        if hasattr(ctx, 'excel_filename') and ctx.excel_filename:
            out_path = self.output_dir / ctx.excel_filename
        else:
            # fallback to title-based filename
            safe_title = ctx.report_title.replace('/', '_')
            out_path = self.output_dir / f"{safe_title}.xlsx"

        # prepare rows generator: include header and metadata lines similar to Google implementation
        def row_generator():
            yield [f"Rapport gemaakt op {ctx.now_utc} met data uit:"]
            yield [f"{ctx.datasource_name}, laatst gesynchroniseerd op {result.last_data_update or ''}"]

            headers = [k.split('.')[-1] for k in result.keys]
            if persistent_column:
                headers = headers + ['opmerkingen (blijvend)']
            yield headers

            persistent_dict_local = persistent_dict or {}
            for out_row in result.iter_rows():
                row = list(out_row)
                if persistent_column:
                    if row and row[0] in persistent_dict_local:
                        row.append(persistent_dict_local[row[0]])
                    else:
                        row.append('')
                yield row

        # If requested, convert specified columns to numbers. The Excel writer writes
        # two metadata rows followed by a header row, then data rows. We wrap the
        # original row generator to perform conversion only on data rows (i.e. rows
        # after the first three yielded rows).
        gen = row_generator()
        if convert_columns_to_numbers:
            try:
                # lazy import to avoid top-level dependency
                from outputs.sheets_cell import SheetsCell

                col_indices = [SheetsCell(f'{col}1')._column_int - 1 for col in convert_columns_to_numbers]

                def _convert_rows(g):
                    row_index = 0
                    for r in g:
                        # r is expected to be a list
                        if not isinstance(r, list):
                            yield r
                            row_index += 1
                            continue

                        # metadata rows (0,1) and header row (2) should not be converted
                        if row_index >= 3:
                            for col_idx in col_indices:
                                if col_idx < len(r):
                                    val = r[col_idx]
                                    if val is not None and val != '' and val != 'None':
                                        try:
                                            # try converting to float; leave as-is on failure
                                            r[col_idx] = float(val)
                                        except Exception:
                                            pass
                        yield r
                        row_index += 1

                gen = _convert_rows(gen)
            except Exception:
                # best-effort: if conversion setup fails, fall back to original generator
                gen = row_generator()

        # write streaming
        meta = self.write_data_to_sheet(out_path, 'Resultaat', gen, start_cell=startcell, overwrite=True)

        # If requested, set an Excel auto-filter on the header row of the Resultaat sheet.
        # We reopen the workbook in normal mode (not write-only) to apply the auto-filter
        # because write-only worksheets do not support setting these properties.
        if add_filter:
            try:
                # clear any existing filter first (compat with Google Sheets behaviour)
                try:
                    self.clear_filter(out_path, 'Resultaat')
                except Exception:
                    # ignore clearing errors; we'll still attempt to set a new filter below
                    pass
                # local imports to avoid hard dependency at module import time
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                from outputs.sheets_cell import SheetsCell

                wb = load_workbook(out_path)
                if 'Resultaat' in wb.sheetnames:
                    ws = wb['Resultaat']

                    # determine start row/column from startcell (e.g. 'A1')
                    sc = SheetsCell(startcell)
                    start_row = sc.row

                    # current Excel write_report writes two metadata rows before the header
                    header_row = start_row + 2

                    # number of header columns = number of keys + optional persistent column
                    num_cols = len(result.keys) + (1 if persistent_column else 0)

                    start_col_index = sc._column_int
                    end_col_index = start_col_index + max(0, num_cols - 1)
                    start_col_letter = get_column_letter(start_col_index)
                    end_col_letter = get_column_letter(end_col_index)

                    # last row: header + number of data rows
                    try:
                        total_data_rows = len(result.rows) if hasattr(result, 'rows') else 0
                    except Exception:
                        total_data_rows = 0
                    last_row = header_row + max(0, total_data_rows)

                    # set the auto-filter range covering header and data rows (matches Google behaviour)
                    ws.auto_filter.ref = f"{start_col_letter}{header_row}:{end_col_letter}{last_row}"

                    # persist changes
                    wb.save(out_path)
            except Exception:
                # best-effort: don't fail the whole report because auto-filter couldn't be set
                pass

            # freeze top rows to match Google Sheets behaviour: freeze rows == header start row
            try:
                from outputs.sheets_cell import SheetsCell
                # compute rows to freeze consistent with header_row above
                sc = SheetsCell(startcell)
                header_row = sc.row + 2
                try:
                    # call helper to freeze top rows (best-effort)
                    self.freeze_top_rows(out_path, 'Resultaat', header_row)
                except Exception:
                    pass
            except Exception:
                pass

            # auto-resize columns to fit content (best-effort)
            try:
                # number of header columns = number of keys + optional persistent column
                num_cols = len(result.keys) + (1 if persistent_column else 0)
                try:
                    self.automatic_resize_columns(out_path, 'Resultaat', number_of_columns=num_cols)
                except Exception:
                    pass
            except Exception:
                pass

            # hyperlink first column (best-effort)
            try:
                # determine first data row start cell
                from outputs.sheets_cell import SheetsCell
                sc = SheetsCell(startcell)
                header_row = sc.row + 2
                data_start_row = header_row + 1
                start_cell_for_links = f"{sc._column_str}{data_start_row}"

                # build first column values from QueryResult.rows
                first_column = []
                try:
                    for r in result.iter_rows():
                        if r:
                            first_column.append(r[0])
                except Exception:
                    # fallback to direct rows list
                    try:
                        for r in getattr(result, 'rows', []):
                            first_column.append(r[0] if r else None)
                    except Exception:
                        first_column = []

                try:
                    self.add_hyperlink_column(out_path, 'Resultaat', start_cell_for_links, link_type, first_column)
                except Exception:
                    pass
            except Exception:
                pass

        # recalc: if requested, ensure formulas are written and force Excel to recalculate on open
        if recalculate_cells:
            try:
                for sheet_name, cell in recalculate_cells:
                    try:
                        self.recalculate_formula(out_path, sheet_name, cell)
                    except Exception:
                        pass
            except Exception:
                pass

        # update summary/historiek is still performed by DQReport via Google Sheets wrapper
        return meta

    def _log_file_change(self, workbook_path: str, action: str) -> None:
        """Helper: append a file change entry by invoking scripts/file_change_log.py.

        Use subprocess to avoid importing scripts as a package. Fail silently on errors.
        """
        try:
            import subprocess, sys
            repo_root = Path(__file__).resolve().parents[1]
            log_script = repo_root / 'scripts' / 'file_change_log.py'
            if not log_script.exists():
                return
            subprocess.run([sys.executable, str(log_script), '--action', action, '--path', str(workbook_path)], check=False)
        except Exception:
            pass

    # --- additional compatibility write helpers --------------------------------
    def clear_filter(self, spreadsheet_id: str | Path, sheet_name: str) -> None:
        """Clear any auto-filter on the given sheet (best-effort).

        Mirrors the Google Sheets wrapper `clear_filter` semantics for Excel files.
        """
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        # Best-effort: remove <autoFilter .../> tags from worksheet XML files inside the .xlsx archive.
        if not wb_path.exists():
            return
        import zipfile, re, os
        from tempfile import NamedTemporaryFile

        try:
            with zipfile.ZipFile(wb_path, 'r') as zin:
                names = zin.namelist()
                files = {name: zin.read(name) for name in names}

            changed = False
            # remove autoFilter tags from worksheet xmls
            for name, data in list(files.items()):
                if name.startswith('xl/worksheets/') and b'<autoFilter' in data:
                    # remove self-closing autoFilter elements
                    new = re.sub(rb'<autoFilter[^>]*/>', b'', data)
                    # also remove opening/closing if not self-closing (defensive)
                    new = re.sub(rb'<autoFilter[^>]*>', b'', new)
                    new = re.sub(rb'</autoFilter>', b'', new)
                    if new != data:
                        files[name] = new
                        changed = True

            if not changed:
                return

            # write out a new zip archive atomically
            dirpath = wb_path.parent
            with NamedTemporaryFile(prefix=wb_path.stem + '_tmp_', suffix=wb_path.suffix, dir=str(dirpath), delete=False) as tmp:
                tmp_name = tmp.name
            try:
                with zipfile.ZipFile(tmp_name, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                    for name, data in files.items():
                        zout.writestr(name, data)
                os.replace(tmp_name, str(wb_path))
            finally:
                try:
                    if os.path.exists(tmp_name):
                        os.remove(tmp_name)
                except Exception:
                    pass
        except Exception:
            # best-effort: ignore failures
            pass

    def create_basic_filter(self, spreadsheet_id: str | Path, sheet_name: str, range: str) -> None:
        """Create an auto-filter on the given sheet covering the specified A1 range.

        This mirrors the Google Sheets `create_basic_filter` helper used by the
        Google writer. Best-effort: if the workbook cannot be opened or the sheet
        is missing, fail silently.
        """
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return
        try:
            wb = load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
            try:
                ws.auto_filter.ref = range
            except Exception:
                # some openpyxl versions may behave differently; set attribute via fallback
                try:
                    setattr(ws, 'auto_filter', None)
                    ws.auto_filter.ref = range
                except Exception:
                    pass
            # persist changes
            try:
                self._atomic_save_workbook(wb, wb_path)
            except Exception:
                try:
                    wb.save(wb_path)
                except Exception:
                    pass
        except Exception:
            # best-effort
            pass

    def freeze_top_rows(self, spreadsheet_id: str | Path, sheet_name: str, rows: int) -> None:
        """Freeze the top `rows` rows on the given sheet (best-effort).

        Mirrors SheetsWrapper.freeze_top_rows for Excel files.
        """
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return
        try:
            wb = load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
            # freeze_panes expects a cell below the frozen rows
            freeze_cell_row = max(1, int(rows) + 1)
            # default to column A
            freeze_cell = f"A{freeze_cell_row}"
            try:
                ws.freeze_panes = ws[freeze_cell]
            except Exception:
                try:
                    ws.freeze_panes = freeze_cell
                except Exception:
                    # ignore
                    pass

            # save atomically
            try:
                self._atomic_save_workbook(wb, wb_path)
            except Exception:
                try:
                    wb.save(wb_path)
                except Exception:
                    pass
        except Exception:
            # best-effort
            pass

    def automatic_resize_columns(self, spreadsheet_id: str | Path, sheet_name: str, number_of_columns: int = 1) -> None:
        """Automatically resize the first `number_of_columns` columns based on content width.

        This is a heuristic: it computes the maximum string length of values in each column
        and sets `ws.column_dimensions[col_letter].width` accordingly. Best-effort.
        """
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return
        try:
            wb = load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]

            from openpyxl.utils import get_column_letter

            # determine bounds
            max_row = ws.max_row or 1
            max_col = min(number_of_columns, ws.max_column or number_of_columns)

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx, values_only=True):
                    try:
                        v = row[0]
                        if v is None:
                            l = 0
                        else:
                            l = len(str(v))
                        if l > max_len:
                            max_len = l
                    except Exception:
                        continue

                # heuristic for width: allow some padding, cap to reasonable max
                if max_len <= 0:
                    width = 8
                else:
                    width = min(100, max(8, int(max_len * 1.1) + 2))

                try:
                    ws.column_dimensions[col_letter].width = width
                except Exception:
                    # ignore
                    pass

            # save changes
            try:
                self._atomic_save_workbook(wb, wb_path)
            except Exception:
                try:
                    wb.save(wb_path)
                except Exception:
                    pass
        except Exception:
            # best-effort
            pass

    def add_hyperlink_column(self, spreadsheet_id: str | Path, sheet_name: str, start_cell: str,
                             link_type: str = 'awvinfra', column_data: list = None) -> None:
        """Add a column of hyperlinks starting at start_cell. column_data is a list of ids/values.

        For each non-empty value in column_data, compute a URL depending on link_type and set
        the cell's hyperlink and value. Best-effort.
        """
        if not column_data:
            return

        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return

        try:
            wb = load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]

            # parse start_cell into column letters and start row
            from outputs.sheets_cell import SheetsCell
            sc = SheetsCell(start_cell)
            col_letter = sc._column_str
            row = sc.row

            for i, val in enumerate(column_data):
                target_row = row + i
                try:
                    cell = ws[f"{col_letter}{target_row}"]
                    if val is None or val == '':
                        cell.value = ''
                        cell.hyperlink = None
                    else:
                        url = ''
                        if link_type == 'awvinfra':
                            url = f'https://apps.mow.vlaanderen.be/awvinfra/ui/#/?asset={val}'
                        elif link_type in {'eminfra', 'eminfra_onderdeel'}:
                            url = f'https://apps.mow.vlaanderen.be/eminfra/assets/{val}'
                        else:
                            # fallback: treat val as url
                            url = str(val)
                        try:
                            cell.hyperlink = url
                        except Exception:
                            try:
                                # older openpyxl may require setting hyperlink as string
                                cell._hyperlink = url
                            except Exception:
                                pass
                        cell.value = str(val)
                except Exception:
                    # ignore per-cell failures
                    continue

            # save
            try:
                self._atomic_save_workbook(wb, wb_path)
            except Exception:
                try:
                    wb.save(wb_path)
                except Exception:
                    pass
        except Exception:
            # best-effort
            pass
        
    def recalculate_formula(self, spreadsheet_id: str | Path, sheet_name: str, cell: str) -> None:
        """Ensure the formula in the given cell is present and instruct Excel to recalc on open.

        This writes the formula if present (openpyxl does not evaluate formulas) and sets
        the workbook properties to force recalculation when the workbook is opened in Excel.
        """
        try:
            _ensure_openpyxl_loaded()
        except Exception:
            return

        wb_path = self._resolve_workbook_path(spreadsheet_id)
        if not wb_path.exists():
            return

        try:
            wb = load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]

            # read raw cell value; if it appears to be a formula (starts with '='), ensure it's set
            try:
                val = ws[cell].value
            except Exception:
                val = None

            # If the cell contains a formula, re-set it (best-effort). If it's empty, nothing to do.
            try:
                if isinstance(val, str) and val.startswith('='):
                    ws[cell].value = val
            except Exception:
                pass

            # Force Excel to recalc on open
            try:
                # newer openpyxl versions expose calc_properties
                try:
                    wb.calc_properties.fullCalcOnLoad = True
                except Exception:
                    try:
                        wb.calculation_properties.fullCalcOnLoad = True
                    except Exception:
                        # last-resort: try setting an attribute
                        try:
                            wb.properties.fullCalcOnLoad = True
                        except Exception:
                            pass
            except Exception:
                pass

            # save workbook
            try:
                self._atomic_save_workbook(wb, wb_path)
            except Exception:
                try:
                    wb.save(wb_path)
                except Exception:
                    pass
        except Exception:
            # best-effort
            pass

        # As a further step, ensure the workbook calcPr/fullCalcOnLoad is present in xl/workbook.xml
        try:
            import zipfile, re, os
            from tempfile import NamedTemporaryFile

            if not wb_path.exists():
                return

            with zipfile.ZipFile(wb_path, 'r') as zin:
                names = zin.namelist()
                if 'xl/workbook.xml' not in names:
                    return
                wb_xml = zin.read('xl/workbook.xml')

            new_xml = wb_xml
            # if calcPr exists, ensure fullCalcOnLoad="1" is present
            if b'<calcPr' in new_xml:
                # add or replace fullCalcOnLoad attribute
                def _add_attr(m):
                    s = m.group(0)
                    if b'fullCalcOnLoad' in s:
                        return s
                    # insert attribute before closing
                    if s.endswith(b'/>'):
                        return s[:-2] + b' fullCalcOnLoad="1"/>'
                    elif s.endswith(b'>'):
                        return s[:-1] + b' fullCalcOnLoad="1">'
                    return s

                new_xml = re.sub(rb'<calcPr[^>]*/?>', _add_attr, new_xml)
            else:
                # insert a self-closing calcPr element after workbookPr if present, else after <workbook>
                m = re.search(rb'<workbookPr[^>]*/?>', new_xml)
                insert_at = None
                if m:
                    insert_at = m.end()
                else:
                    m2 = re.search(rb'<workbook[^>]*>', new_xml)
                    insert_at = m2.end() if m2 else None

                if insert_at is not None:
                    insertion = b'<calcPr fullCalcOnLoad="1"/>'
                    new_xml = new_xml[:insert_at] + insertion + new_xml[insert_at:]

            if new_xml != wb_xml:
                # write out a new zip archive atomically
                dirpath = wb_path.parent
                with NamedTemporaryFile(prefix=wb_path.stem + '_tmp_', suffix=wb_path.suffix, dir=str(dirpath), delete=False) as tmp:
                    tmp_name = tmp.name
                try:
                    with zipfile.ZipFile(wb_path, 'r') as zin:
                        with zipfile.ZipFile(tmp_name, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                            for item in zin.infolist():
                                if item.filename == 'xl/workbook.xml':
                                    zout.writestr(item, new_xml)
                                else:
                                    zout.writestr(item, zin.read(item.filename))
                    os.replace(tmp_name, str(wb_path))
                finally:
                    try:
                        if os.path.exists(tmp_name):
                            os.remove(tmp_name)
                    except Exception:
                        pass
        except Exception:
            # best-effort
            pass


# Export
__all__ = ["ExcelOutput"]
