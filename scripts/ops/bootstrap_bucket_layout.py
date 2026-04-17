#!/usr/bin/env python3
"""One-shot bootstrap for a reset RSA_OneDrive folder.

What it does:
- discovers the configured RSA reports
- creates the expected 100-report bucket folders
- moves the current report workbooks into their bucket
- prefers the more complete workbook when both root and bucket copies exist
- updates the local summary workbook hyperlinks so they point to bucketed SharePoint URLs

Usage:
  uv run python -m scripts.ops.bootstrap_bucket_layout --output-dir RSA_OneDrive
  uv run python -m scripts.ops.bootstrap_bucket_layout --dry-run
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass
from contextlib import suppress
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

# Ensure repository root is on sys.path so imports work when executed from other cwd
repo_root = Path(__file__).resolve().parents[2]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

from lib.reports.instantiator import discover_and_instantiate_reports
from outputs.excel import ExcelOutput
from outputs.report_routes import extract_report_number, report_bucket_name, report_sharepoint_url, resolve_report_output_path
from outputs.spreadsheet_map import lookup as lookup_spreadsheet_filename

DEFAULT_OUTPUT_DIR = repo_root / 'RSA_OneDrive'
DEFAULT_OVERVIEW_DIR = 'Overzicht'
DEFAULT_SUMMARY_NAME = '[RSA] Overzicht rapporten.xlsx'
BRUSSELS_SHAREPOINT_BASE = 'https://vlaamseoverheid.sharepoint.com/sites/AIW_AIM_BIM/AIM_BIM/RSA'

logging.basicConfig(level=logging.INFO, format='[BOOT] %(message)s')
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ReportDescriptor:
    report_name: str
    report_title: str
    excel_filename: str
    report_number: Optional[int]


@dataclass
class BootstrapStats:
    discovered: int = 0
    bucket_dirs_created: int = 0
    reports_migrated: int = 0
    reports_replaced: int = 0
    reports_kept: int = 0
    summary_rows_updated: int = 0
    stale_locks_removed: int = 0
    missing_reports: int = 0


def _make_descriptor(instance) -> Optional[ReportDescriptor]:
    try:
        if hasattr(instance, 'init_report'):
            instance.init_report()
    except Exception as exc:
        logger.warning('Skipping %s: init_report failed (%s)', type(instance).__name__, exc)
        return None

    report = getattr(instance, 'report', None)
    if report is None:
        logger.warning('Skipping %s: init_report did not set self.report', type(instance).__name__)
        return None

    report_name = str(getattr(report, 'name', '') or type(instance).__name__)
    report_title = str(getattr(report, 'title', '') or '')
    excel_filename = str(getattr(report, 'excel_filename', '') or '').strip()
    if not excel_filename:
        excel_filename = str(lookup_spreadsheet_filename(getattr(report, 'spreadsheet_id', '') or '') or '').strip()
    if not excel_filename:
        logger.warning('Skipping %s: no excel_filename configured', report_name)
        return None

    report_number = extract_report_number(report_name, report_title, excel_filename)
    return ReportDescriptor(
        report_name=report_name,
        report_title=report_title,
        excel_filename=excel_filename,
        report_number=report_number,
    )


def discover_report_descriptors() -> list[ReportDescriptor]:
    descriptors: list[ReportDescriptor] = []
    for instance in discover_and_instantiate_reports():
        desc = _make_descriptor(instance)
        if desc is not None:
            descriptors.append(desc)
    return descriptors


def _is_valid_workbook(path: Path) -> bool:
    if not path.exists() or not path.is_file():
        return False
    if path.suffix.lower() != '.xlsx':
        return False
    if path.name.endswith('.lock') or '_tmp_' in path.name:
        return False
    try:
        wb = load_workbook(path, read_only=True)
        try:
            return bool(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return False


def _workbook_score(path: Path) -> int:
    if not _is_valid_workbook(path):
        try:
            return path.stat().st_size
        except Exception:
            return -1

    score = 0
    try:
        score += path.stat().st_size
    except Exception:
        pass

    try:
        wb = load_workbook(path, read_only=True)
        try:
            score += len(wb.sheetnames) * 1_000_000
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    score += (ws.max_row or 0) * 10_000
                    score += (ws.max_column or 0) * 1_000
                    if sheet_name in {'Overzicht', 'Historiek'}:
                        score += 50_000_000
                except Exception:
                    continue
        finally:
            wb.close()
    except Exception:
        pass

    return score


def _cleanup_stale_lock(path: Path, dry_run: bool = False, stats: BootstrapStats | None = None) -> bool:
    lock_path = Path(path).with_suffix(f'{path.suffix}.lock')
    if not lock_path.exists():
        return False
    if dry_run:
        logger.info('Would remove stale lock: %s', lock_path)
        return True
    try:
        with suppress(Exception):
            lock_path.unlink()
        logger.info('Removed stale lock: %s', lock_path)
        if stats is not None:
            stats.stale_locks_removed += 1
        return True
    except Exception as exc:
        logger.warning('Could not remove stale lock %s (%s)', lock_path, exc)
        return False


def _move_or_keep_best(source: Path, target: Path, *, dry_run: bool = False, stats: BootstrapStats | None = None) -> str:
    source_exists = source.exists()
    target_exists = target.exists()

    if not source_exists and not target_exists:
        return 'missing'

    if source == target:
        return 'already_bucketed'

    if source_exists and not target_exists:
        if dry_run:
            logger.info('Would move %s -> %s', source, target)
        else:
            target.parent.mkdir(parents=True, exist_ok=True)
            source.replace(target)
        _cleanup_stale_lock(source, dry_run=dry_run, stats=stats)
        _cleanup_stale_lock(target, dry_run=dry_run, stats=stats)
        return 'moved'

    if target_exists and not source_exists:
        _cleanup_stale_lock(target, dry_run=dry_run, stats=stats)
        return 'kept'

    # Both exist: keep the more complete workbook.
    source_score = _workbook_score(source)
    target_score = _workbook_score(target)

    if source_score > target_score:
        if dry_run:
            logger.info('Would replace bucket copy %s with more complete %s (scores %s > %s)', target, source, source_score, target_score)
        else:
            target.unlink(missing_ok=True)
            source.replace(target)
        _cleanup_stale_lock(source, dry_run=dry_run, stats=stats)
        _cleanup_stale_lock(target, dry_run=dry_run, stats=stats)
        return 'replaced'

    if dry_run:
        logger.info('Would remove redundant source %s (bucket copy kept, scores %s >= %s)', source, target_score, source_score)
    else:
        source.unlink(missing_ok=True)
    _cleanup_stale_lock(source, dry_run=dry_run, stats=stats)
    _cleanup_stale_lock(target, dry_run=dry_run, stats=stats)
    return 'kept_bucket'


def _find_summary_row(ws, descriptor: ReportDescriptor) -> Optional[int]:
    target_name = descriptor.report_name.strip().lower()
    target_title = descriptor.report_title.strip().lower()
    target_file = descriptor.excel_filename.strip().lower()

    for row_idx in range(4, (ws.max_row or 0) + 1):
        f_value = ws.cell(row=row_idx, column=6).value
        if f_value is not None and str(f_value).strip().lower() == target_name:
            return row_idx

    for row_idx in range(4, (ws.max_row or 0) + 1):
        b_cell = ws.cell(row=row_idx, column=2)
        candidates = [b_cell.value, getattr(b_cell, 'hyperlink', None)]
        for candidate in candidates:
            if not candidate:
                continue
            text = str(candidate).strip().lower()
            if target_file and target_file in text:
                return row_idx
            if target_name and target_name in text:
                return row_idx
            if target_title and target_title in text:
                return row_idx

    return None


def _update_summary_workbook(summary_path: Path, descriptors: Iterable[ReportDescriptor], *, dry_run: bool = False) -> int:
    if not summary_path.exists():
        logger.warning('Summary workbook not found: %s', summary_path)
        return 0

    wb = load_workbook(summary_path)
    try:
        if 'Overzicht' not in wb.sheetnames:
            logger.warning('Summary workbook has no Overzicht sheet: %s', summary_path)
            return 0

        ws = wb['Overzicht']
        updated = 0
        for descriptor in descriptors:
            row_idx = _find_summary_row(ws, descriptor)
            if row_idx is None:
                logger.warning('Could not find summary row for %s', descriptor.report_name)
                continue

            url = report_sharepoint_url(
                excel_filename=descriptor.excel_filename,
                report_name=descriptor.report_name,
                report_title=descriptor.report_title,
                base_url=BRUSSELS_SHAREPOINT_BASE,
            )
            b_cell = ws.cell(row=row_idx, column=2)
            if url:
                b_cell.value = 'Link'
                b_cell.hyperlink = url
                try:
                    b_cell.style = 'Hyperlink'
                except Exception:
                    pass
            else:
                b_cell.value = None
                try:
                    b_cell.hyperlink = None
                except Exception:
                    pass
            updated += 1

        if not dry_run:
            writer = ExcelOutput(output_dir=str(summary_path.parent))
            writer._atomic_save_workbook(wb, summary_path)
        return updated
    finally:
        try:
            wb.close()
        except Exception:
            pass


def bootstrap_bucket_layout(output_dir: Path, *, dry_run: bool = False) -> tuple[BootstrapStats, list[ReportDescriptor]]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    stats = BootstrapStats()
    descriptors = discover_report_descriptors()
    stats.discovered = len(descriptors)

    bucket_names: set[str] = set()
    for descriptor in descriptors:
        if descriptor.report_number is not None:
            bucket_names.add(report_bucket_name(descriptor.report_number))

    for bucket in sorted(bucket_names):
        bucket_dir = output_dir / bucket
        if bucket_dir.exists():
            continue
        stats.bucket_dirs_created += 1
        if dry_run:
            logger.info('Would create bucket directory: %s', bucket_dir)
        else:
            bucket_dir.mkdir(parents=True, exist_ok=True)
            logger.info('Created bucket directory: %s', bucket_dir)

    for descriptor in descriptors:
        if descriptor.report_number is None:
            logger.warning('Skipping %s: no report number could be extracted', descriptor.report_name)
            stats.missing_reports += 1
            continue

        target = resolve_report_output_path(
            output_dir,
            excel_filename=descriptor.excel_filename,
            report_name=descriptor.report_name,
            report_title=descriptor.report_title,
        )
        source = output_dir / Path(descriptor.excel_filename).name
        result = _move_or_keep_best(source, target, dry_run=dry_run, stats=stats)
        if result == 'missing':
            logger.warning('Missing report workbook: %s', source)
            stats.missing_reports += 1
        elif result in {'moved', 'replaced'}:
            stats.reports_migrated += 1
            if result == 'replaced':
                stats.reports_replaced += 1
        elif result in {'kept', 'kept_bucket', 'already_bucketed'}:
            stats.reports_kept += 1

    summary_path = output_dir / DEFAULT_OVERVIEW_DIR / DEFAULT_SUMMARY_NAME
    stats.summary_rows_updated = _update_summary_workbook(summary_path, descriptors, dry_run=dry_run)

    return stats, descriptors


def main() -> int:
    parser = argparse.ArgumentParser(description='Bootstrap RSA_OneDrive into bucket layout and refresh summary links')
    parser.add_argument('--output-dir', default=str(DEFAULT_OUTPUT_DIR), help='Path to RSA_OneDrive')
    parser.add_argument('--dry-run', action='store_true', help='Show actions without modifying files')
    args = parser.parse_args()

    stats, descriptors = bootstrap_bucket_layout(Path(args.output_dir), dry_run=args.dry_run)

    print('Bootstrap finished')
    print(f'  reports discovered : {stats.discovered}')
    print(f'  bucket dirs created : {stats.bucket_dirs_created}')
    print(f'  reports migrated    : {stats.reports_migrated}')
    print(f'  reports replaced    : {stats.reports_replaced}')
    print(f'  reports kept        : {stats.reports_kept}')
    print(f'  summary rows updated: {stats.summary_rows_updated}')
    print(f'  stale locks removed : {stats.stale_locks_removed}')
    print(f'  missing reports     : {stats.missing_reports}')

    return 0 if descriptors else 1


if __name__ == '__main__':
    raise SystemExit(main())

