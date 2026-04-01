"""Aggregator for staged summary JSON files.

Scans a staging directory, claims files (moves to processing/), applies operations using ExcelOutput,
and moves processed files to processed/ or failed/.

Usage (dry-run):
  python scripts/ops/aggregate_summaries.py --staged-dir RSA_OneDrive/staged_summaries --dry-run

Usage (real run):
  python scripts/ops/aggregate_summaries.py --staged-dir RSA_OneDrive/staged_summaries --output-dir RSA_OneDrive --limit 100
"""
from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path
from typing import Any, Dict
from datetime import datetime
from zoneinfo import ZoneInfo
import re
import sys

repo_root = Path(__file__).resolve().parents[2]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

from outputs.excel import ExcelOutput
from outputs.summary_stager import _ensure_dirs
from openpyxl import load_workbook
from pathlib import Path as _Path

logging.basicConfig(level=logging.INFO, format='[AGG] %(asctime)s %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
BRUSSELS = ZoneInfo('Europe/Brussels')


def apply_payload(excel: ExcelOutput, payload: Dict[str, Any], output_dir: Path) -> Path | None:
    op = payload['operation']
    fname = payload.get('excel_filename') or payload.get('spreadsheet_id')
    sheet = payload['sheet']

    # If payload contains a report name, prefer locating the workbook by scanning
    # all workbooks for the report entry in Overzicht column F (report number).
    report_name = payload.get('meta', {}).get('report')
    wb_path = None
    # Prefer the canonical summary workbook if present when updating the Overzicht sheet
    try:
        # Use the ExcelOutput's configured output_dir (absolute) so resolution
        # matches ExcelOutput._resolve_workbook_path and any mappings discovered
        # earlier when grouping writes. This avoids inconsistencies when a
        # relative Path(output_dir) differs from the ExcelOutput instance's
        # resolved output directory and prevents selecting the wrong workbook
        # (e.g. 'ABBAMelda - EM-Infra id_s.xlsx') when the canonical
        # '[RSA] Overzicht rapporten.xlsx' is present in the ExcelOutput folder.
        preferred_summary = Path(getattr(excel, 'output_dir', Path(output_dir))) / '[RSA] Overzicht rapporten.xlsx'
        if payload.get('sheet') == 'Overzicht' and preferred_summary.exists():
            wb_path = preferred_summary.resolve()
            logger.debug('Using explicit summary workbook %s for Overzicht writes', wb_path)
    except Exception:
        logger.debug('Failed to check preferred summary workbook', exc_info=True)

    # Fallback: resolve workbook path using the ExcelOutput resolver (mapping / filename rules)
    if wb_path is None:
        try:
            wb_path = excel._resolve_workbook_path(fname)
            logger.debug('Resolved workbook path for "%s" -> %s (exists=%s)', fname, wb_path, wb_path.exists())
        except Exception:
            logger.debug('Failed to resolve workbook path for %s', fname, exc_info=True)

    if op == 'append_row':
        row = payload['row']
        # append by reading existing rows and writing back with new appended row via write_data_to_sheet
        existing = excel.read_data_from_sheet(str(wb_path), sheet) or []
        # Preserve header: if sheet has header, insert new row after header (index 0 -> header)
        # Determine if first row looks like header (simple heuristic: values are strings and not UUIDs)
        if len(existing) > 0:
            # insert after header: keep existing[0] as header, new row becomes position 1
            combined = [existing[0]] + [row] + existing[1:]
        else:
            combined = [row]
        excel.write_data_to_sheet(wb_path, sheet, combined, overwrite=True)
        # If this is the Historiek sheet (DQReport writes), apply automatic column
        # resizing so the newly appended data is visible without manual adjustment.
        try:
            # determine number of columns from header if present else from the new row
            num_cols = 0
            if combined and isinstance(combined, list) and len(combined) > 0 and isinstance(combined[0], (list, tuple)):
                num_cols = len(combined[0])
            elif isinstance(row, (list, tuple)):
                num_cols = len(row)
            if num_cols > 0:
                try:
                    excel.automatic_resize_columns(wb_path, sheet, number_of_columns=num_cols)
                except Exception:
                    logger.exception('Failed automatic resize for %s sheet %s', wb_path, sheet)
        except Exception:
            logger.exception('Failed to schedule automatic resize for %s sheet %s', wb_path, sheet)
        return wb_path.resolve()

    elif op == 'write_cell':
        cell = payload['cell']
        value = payload['value']
        wb_path = Path(wb_path).resolve()
        logger.debug('Applying write_cell to workbook %s sheet %s cell %s (report=%s)', wb_path, sheet, cell, payload.get('meta',{}).get('report'))
        # If writing to Overzicht column C, normalize the timestamp format to Brussels 'YYYY-MM-DD HH:MM:SS'
        def _normalize_to_brussels_string(val) -> str:
            if val is None:
                return ''
            if isinstance(val, (int, float)):
                # numeric values: return as-is
                return str(val)
            s = str(val).strip()
            if s == '':
                return ''
            # try to parse ISO format first
            try:
                # Python's fromisoformat handles many ISO variants but not trailing Z; replace Z with +00:00
                iso = s.replace('Z', '+00:00') if s.endswith('Z') else s
                dt = None
                try:
                    dt = datetime.fromisoformat(iso)
                except Exception:
                    dt = None
                if dt is None:
                    # try common formats
                    fmts = [
                        '%Y-%m-%d %H:%M:%S',
                        '%Y-%m-%dT%H:%M:%S.%f%z',
                        '%Y-%m-%dT%H:%M:%S%z',
                        '%Y-%m-%d %H:%M:%S%z',
                        '%Y-%m-%dT%H:%M:%S.%f',
                        '%Y-%m-%dT%H:%M:%S',
                    ]
                    for f in fmts:
                        try:
                            dt = datetime.strptime(s, f)
                            break
                        except Exception:
                            dt = None
                if dt is None:
                    # last-resort: extract leading datetime-like substring
                    m = re.search(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[+-]\d{2}:?\d{2}|Z)?", s)
                    if m:
                        piece = m.group(0).replace('Z', '+00:00')
                        try:
                            dt = datetime.fromisoformat(piece)
                        except Exception:
                            dt = None
                if dt is None:
                    return s
                # ensure tzinfo: if naive, assume Brussels (user requested Brussels) and convert to Brussels
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=BRUSSELS)
                # convert to Brussels and format
                dt_utc = dt.astimezone(BRUSSELS)
                return dt_utc.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return str(val)

        def _parse_dt(val):
            """Return a timezone-aware datetime in Brussels if parseable, else None."""
            if val is None:
                return None
            if isinstance(val, datetime):
                dt = val
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=BRUSSELS)
                return dt.astimezone(BRUSSELS)
            s = str(val).strip()
            if s == '':
                return None
            iso = s.replace('Z', '+00:00') if s.endswith('Z') else s
            try:
                dt = datetime.fromisoformat(iso)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=BRUSSELS)
                return dt.astimezone(BRUSSELS)
            except Exception:
                pass
            # try known formats
            fmts = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S.%f%z',
                '%Y-%m-%dT%H:%M:%S%z',
                '%Y-%m-%d %H:%M:%S%z',
                '%Y-%m-%dT%H:%M:%S.%f',
                '%Y-%m-%dT%H:%M:%S',
            ]
            for f in fmts:
                try:
                    dt = datetime.strptime(s, f)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=BRUSSELS)
                    return dt.astimezone(BRUSSELS)
                except Exception:
                    continue
            # regex fallback
            m = re.search(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[+-]\d{2}:?\d{2}|Z)?", s)
            if m:
                piece = m.group(0).replace('Z', '+00:00')
                try:
                    dt = datetime.fromisoformat(piece)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=BRUSSELS)
                    return dt.astimezone(BRUSSELS)
                except Exception:
                    return None
            return None

        # detect column letter from cell
        col_letters = ''.join([c for c in cell if c.isalpha()])
        if sheet == 'Overzicht' and col_letters.upper() == 'C':
            # value expected to be [timestamp, count]
            # Handle list payloads like [timestamp, count]
            if isinstance(value, list) and len(value) >= 1:
                # Normalize payload timestamp and always write the payload's timestamp (normalized to Brussels).
                try:
                    normalized = _normalize_to_brussels_string(value[0])
                    value[0] = normalized
                    logger.info('Overzicht C write: report=%s workbook=%s cell=%s payload_normalized=%s',
                                payload.get('meta', {}).get('report'), str(Path(wb_path).resolve()), cell, normalized)
                except Exception:
                    pass
            else:
                # Handle scalar timestamp payloads
                try:
                    normalized = _normalize_to_brussels_string(value)
                    # write back normalized scalar
                    before_val = None
                    try:
                        from openpyxl import load_workbook
                        wb_existing = load_workbook(wb_path, read_only=True)
                        ws_existing = wb_existing['Overzicht']
                        row_idx = int(''.join([c for c in cell if c.isdigit()]))
                        before_val = ws_existing.cell(row=row_idx, column=3).value
                    except Exception:
                        before_val = None
                    value = normalized
                    logger.info('Overzicht C write normalize (scalar): report=%s workbook=%s cell=%s before=%s after=%s',
                                payload.get('meta', {}).get('report'), str(Path(wb_path).resolve()), cell, before_val, normalized)
                except Exception:
                    pass

        # If value is a list, write horizontally across columns starting at `cell`.
        if isinstance(value, list):
            # compute start column and row
            from outputs.sheets_cell import SheetsCell
            sc = SheetsCell(cell)
            col_index = sc._column_int
            row_index = sc._row
            # Verify that the target row contains the expected report name in column F.
            report_name = (payload.get('meta', {}).get('report') or '').strip().lower()
            try:
                if report_name:
                    from openpyxl import load_workbook as _load_wb
                    _wb_check = _load_wb(wb_path, read_only=True)
                    if sheet in _wb_check.sheetnames:
                        _ws_check = _wb_check[sheet]
                        existing_f = _ws_check.cell(row=row_index, column=6).value
                        if not (existing_f and str(existing_f).strip().lower() == report_name):
                            # search column F for the report_name
                            found_row = None
                            max_search = min(2000, _ws_check.max_row or 1000)
                            for r in range(4, max_search + 1):
                                try:
                                    fv = _ws_check.cell(row=r, column=6).value
                                except Exception:
                                    fv = None
                                if fv and str(fv).strip().lower() == report_name:
                                    found_row = r
                                    break
                            if found_row:
                                old = row_index
                                row_index = found_row
                                logger.warning('Report name mismatch in %s: staged cell %s pointed at row %s (F=%r); correcting to row %s', wb_path, cell, old, existing_f, row_index)
                                # update SheetsCell/target variables accordingly
                                # col_index remains same; target cell will be recomputed below
                            else:
                                logger.warning('Could not find report %s in column F of %s; leaving staged target %s', report_name, wb_path, cell)
                    _wb_check.close()
            except Exception:
                logger.exception('Failed to verify/locate report row in %s for report=%s', wb_path, report_name)
            # write each element to successive columns
            for i, v in enumerate(value):
                target_col = SheetsCell._convert_number_to_column(col_index + i)
                target_cell = f"{target_col}{row_index}"
                excel.write_single_cell(wb_path, sheet, target_cell, v)
            # Post-write verification log: read back the value written to the first column
            try:
                from openpyxl import load_workbook as _load_wb
                _wb = _load_wb(wb_path, read_only=True)
                if sheet in _wb.sheetnames:
                    _ws = _wb[sheet]
                    new_val = _ws.cell(row=row_index, column=col_index).value
                    logger.info('Post-write Overzicht C check (list): report=%s workbook=%s cell=%s after=%s',
                                payload.get('meta', {}).get('report'), str(Path(wb_path).resolve()), f"{SheetsCell._convert_number_to_column(col_index)}{row_index}", new_val)
            except Exception:
                logger.exception('Failed post-write check for %s %s', wb_path, cell)
            return wb_path.resolve()
        else:
            # scalar
            excel.write_single_cell(wb_path, sheet, cell, value)
            # Post-write verification log for Overzicht C scalar writes
            try:
                from openpyxl import load_workbook as _load_wb
                _wb = _load_wb(wb_path, read_only=True)
                if sheet in _wb.sheetnames:
                    _ws = _wb[sheet]
                    row_idx = int(''.join([c for c in cell if c.isdigit()]))
                    new_val = _ws.cell(row=row_idx, column=3).value
                    logger.info('Post-write Overzicht C check (scalar): report=%s workbook=%s cell=%s after=%s',
                                payload.get('meta', {}).get('report'), str(Path(wb_path).resolve()), cell, new_val)
            except Exception:
                logger.exception('Failed post-write check for %s %s', wb_path, cell)
            return wb_path.resolve()

    elif op == 'increment_cell':
        cell = payload['cell']
        delta = int(payload.get('delta', 1))
        excel.update_row_by_adding_number(wb_path, sheet, cell, delta)
        return wb_path

    else:
        raise ValueError(f'Unknown operation: {op}')

    return None


def process_once(staged_dir: Path, output_dir: Path, limit: int = 100, dry_run: bool = True) -> int:
    _ensure_dirs(staged_dir)
    files = sorted([p for p in staged_dir.iterdir() if p.is_file() and p.suffix == '.json'])
    if not files:
        logger.info('No staged files found in %s', staged_dir)
        return 0

    excel = ExcelOutput(output_dir=str(output_dir))

    # First, read all payloads (up to limit) and group write_cell operations that target
    # Overzicht column C by their target workbook/sheet/cell. For those groups, we will
    # consolidate timestamps and perform a single write to avoid races where an older
    # staged file overwrites a newer timestamp.
    to_process = []
    for f in files[:limit]:
        try:
            with open(f, 'r', encoding='utf-8') as fh:
                payload = json.load(fh)
            to_process.append((f, payload))
        except Exception:
            logger.exception('Failed to read staged file %s', f)

    # Group write_cell Overzicht C payloads
    grouped = {}
    others = []

    def _parse_dt_local(val):
        # lightweight parser: reuse logic similar to _parse_dt in apply_payload
        if val is None:
            return None
        if isinstance(val, datetime):
            dt = val
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=BRUSSELS)
            return dt.astimezone(BRUSSELS)
        s = str(val).strip()
        if s == '':
            return None
        iso = s.replace('Z', '+00:00') if s.endswith('Z') else s
        try:
            dt = datetime.fromisoformat(iso)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=BRUSSELS)
            return dt.astimezone(BRUSSELS)
        except Exception:
            pass
        fmts = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S.%f%z',
            '%Y-%m-%dT%H:%M:%S%z',
            '%Y-%m-%d %H:%M:%S%z',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M:%S',
        ]
        for ffmt in fmts:
            try:
                dt = datetime.strptime(s, ffmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=BRUSSELS)
                return dt.astimezone(BRUSSELS)
            except Exception:
                continue
        m = re.search(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[+-]\d{2}:?\d{2}|Z)?", s)
        if m:
            piece = m.group(0).replace('Z', '+00:00')
            try:
                dt = datetime.fromisoformat(piece)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=BRUSSELS)
                return dt.astimezone(BRUSSELS)
            except Exception:
                return None
        return None

    for f, payload in to_process:
        try:
            if payload.get('operation') == 'write_cell' and payload.get('sheet') == 'Overzicht':
                cell = payload.get('cell', '')
                if cell and cell.upper().startswith('C'):
                    fname = payload.get('excel_filename') or payload.get('spreadsheet_id')
                    wb_path = excel._resolve_workbook_path(fname)
                    key = (str(wb_path), payload.get('sheet'), cell.upper())
                    grouped.setdefault(key, []).append((f, payload))
                    continue
            others.append((f, payload))
        except Exception:
            others.append((f, payload))

    processed = 0
    modified_workbooks = set()

    # Apply grouped consolidated writes first (one write per target cell)
    for key, items in grouped.items():
        wb_path_str, sheet_name, cell = key
        # pick latest timestamp among payloads
        latest_dt = None
        latest_payload = None
        for f, payload in items:
            val = payload.get('value')
            candidate = None
            if isinstance(val, list) and len(val) > 0:
                candidate = _parse_dt_local(val[0])
            else:
                candidate = _parse_dt_local(val)
            if candidate and (latest_dt is None or candidate > latest_dt):
                latest_dt = candidate
                latest_payload = payload
            elif latest_payload is None:
                latest_payload = payload

        # If we found a latest datetime, normalize it into the payload to write
        if latest_payload is not None:
            try:
                # ensure normalized form using apply_payload's normalization via calling helper
                # reuse apply_payload by calling it with a modified payload
                if isinstance(latest_payload.get('value'), list) and len(latest_payload.get('value')) > 0:
                    if latest_dt is not None:
                        latest_payload['value'][0] = latest_dt.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        # fallback normalize string
                        try:
                            latest_payload['value'][0] = str(latest_payload['value'][0])
                        except Exception:
                            pass
                else:
                    if latest_dt is not None:
                        latest_payload['value'] = latest_dt.strftime('%Y-%m-%d %H:%M:%S')
                # apply the consolidated payload
                if dry_run:
                    logger.debug('Dry-run consolidated apply %s', latest_payload)
                else:
                    wb_path = apply_payload(excel, latest_payload, output_dir)
                    if wb_path:
                        modified_workbooks.add(wb_path)
                # move all source files for this key to processed
                for f, _ in items:
                    try:
                        processing = staged_dir / 'processing' / f.name
                        f.replace(processing)
                        processed_path = staged_dir / 'processed' / processing.name
                        processing.replace(processed_path)
                    except Exception:
                        logger.exception('Failed moving grouped file %s to processed', f)
                processed += len(items)
            except Exception:
                logger.exception('Failed to apply consolidated write for %s', key)

    # Now process other payloads normally
    for f, payload in others:
        try:
            processing = staged_dir / 'processing' / f.name
            f.replace(processing)
            logger.debug('Processing %s', processing)
            if dry_run:
                logger.debug('Dry-run: would apply %s', payload)
                processed_path = staged_dir / 'processed' / processing.name
                processing.replace(processed_path)
                logger.debug('Dry-run: moved to processed %s', processed_path)
            else:
                wb_path = apply_payload(excel, payload, output_dir)
                processed_path = staged_dir / 'processed' / processing.name
                processing.replace(processed_path)
                logger.debug('Applied and moved to processed %s', processed_path)
                if wb_path:
                    modified_workbooks.add(wb_path)
            processed += 1
        except Exception as e:
            logger.exception('Failed to process %s: %s', f, e)
            try:
                failed_path = staged_dir / 'failed' / f.name
                f.replace(failed_path)
            except Exception:
                logger.exception('Failed to move to failed for %s', f)

    # Log modified workbooks
    if modified_workbooks:
        logger.info('Modified workbooks:')
        for wb in modified_workbooks:
            logger.info(' - %s', wb)

    return processed


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--staged-dir', default='RSA_OneDrive/staged_summaries')
    p.add_argument('--output-dir', default='RSA_OneDrive')
    p.add_argument('--limit', type=int, default=100)
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()

    staged_dir = Path(args.staged_dir)
    output_dir = Path(args.output_dir)

    logger.info('Scanning staged dir %s', staged_dir)
    n = process_once(staged_dir, output_dir, limit=args.limit, dry_run=args.dry_run)
    logger.info('Processed %d staged files', n)


if __name__ == '__main__':
    main()


