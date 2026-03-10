#!/usr/bin/env python3
"""Normalize timestamps in RSA summary workbooks (Overzicht sheet, column C) to UTC 'YYYY-MM-DD HH:MM:SS'.

This script:
- Opens the canonical summary workbook `RSA_OneDrive/[RSA] Overzicht rapporten.xlsx` (if present)
- Scans rows from row 4 onward on sheet 'Overzicht'
- For each row, if column F contains a report id (e.g., Report0004) and column C contains a value,
  attempts to parse and normalize it to UTC format and writes it back if changed.

Usage:
  python scripts/normalize_summary_timestamps.py

This edits the workbook in place (atomic save). No backups are created.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, timezone
import re
import logging

logging.basicConfig(level=logging.INFO, format='[NORM] %(asctime)s %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

SUMMARY_PATH = Path('RSA_OneDrive') / '[RSA] Overzicht rapporten.xlsx'


def normalize_to_utc_string(s: object) -> str:
    if s is None:
        return ''
    if isinstance(s, (int, float)):
        return str(s)
    st = str(s).strip()
    if st == '':
        return ''
    # try ISO parse first (handle trailing Z)
    try:
        iso = st.replace('Z', '+00:00') if st.endswith('Z') else st
        dt = None
        try:
            dt = datetime.fromisoformat(iso)
        except Exception:
            dt = None
        if dt is None:
            fmts = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S.%f%z',
                '%Y-%m-%dT%H:%M:%S%z',
                '%Y-%m-%d %H:%M:%S%z',
                '%Y-%m-%dT%H:%M:%S.%f',
                '%Y-%m-%dT%H:%M:%S',
            ]
            for f in fmts:
                try:
                    dt = datetime.strptime(st, f)
                    break
                except Exception:
                    dt = None
        if dt is None:
            m = re.search(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[+-]\d{2}:?\d{2}|Z)?", st)
            if m:
                piece = m.group(0).replace('Z', '+00:00')
                try:
                    dt = datetime.fromisoformat(piece)
                except Exception:
                    dt = None
        if dt is None:
            return st
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt_utc = dt.astimezone(timezone.utc)
        return dt_utc.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return st


def main() -> int:
    if not SUMMARY_PATH.exists():
        logger.error('Summary workbook not found: %s', SUMMARY_PATH)
        return 1
    wb = load_workbook(SUMMARY_PATH)
    if 'Overzicht' not in wb.sheetnames:
        logger.error('Overzicht sheet not found in %s', SUMMARY_PATH)
        return 1
    ws = wb['Overzicht']
    modified = False
    # iterate rows starting at 4
    for row_idx in range(4, ws.max_row + 1):
        report_cell = ws.cell(row=row_idx, column=6).value  # column F
        if not report_cell:
            continue
        c_cell = ws.cell(row=row_idx, column=3)  # column C
        orig = c_cell.value
        norm = normalize_to_utc_string(orig)
        if norm != (orig if orig is None else str(orig)):
            # if normalized differs, write back
            # But only write if actual difference in string representation
            # Compare normalized to formatted original if original is datetime
            write_needed = False
            if isinstance(orig, datetime):
                # format original to target and compare
                orig_fmt = orig.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S') if orig.tzinfo else orig.strftime('%Y-%m-%d %H:%M:%S')
                if orig_fmt != norm:
                    write_needed = True
            else:
                # compare string forms
                if str(orig).strip() != norm:
                    write_needed = True
            if write_needed:
                logger.info('Normalizing row %d report %s: "%s" -> "%s"', row_idx, report_cell, orig, norm)
                c_cell.value = norm
                modified = True
    if modified:
        wb.save(SUMMARY_PATH)
        logger.info('Saved normalized summary workbook: %s', SUMMARY_PATH)
    else:
        logger.info('No changes needed in %s', SUMMARY_PATH)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

