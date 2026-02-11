#!/usr/bin/env python3
"""Test harness: run Report0002 but stub out datasources and Google Sheets wrapper so we only exercise Excel output.

Usage:
    python3 scripts/test_report0002_excel.py

This will write an Excel file to RSA_OneDrive as defined in the report's excel_filename.
"""
from pathlib import Path
import importlib
import datetime
import logging

# Ensure project path imports
from Reports.Report0002 import Report0002
import factories
from outputs.excel import ExcelOutput
from datasources.base import QueryResult
from outputs.base import OutputWriteContext

# --- Stubs ---------------------------------------------------------------
class DummyDatasource:
    name = 'Dummy'
    def test_connection(self):
        # pretend OK
        return None
    def execute(self, query: str) -> QueryResult:
        # Return a small deterministic result set matching Report0002 columns (uuid, naam)
        now = datetime.datetime.now(datetime.timezone.utc).isoformat()
        rows = [
            {'uuid': '628edbc5-5c33-4365-8fcf-2970e326f3e4', 'naam': 'tlc-fi-broker'},
            {'uuid': '8c01c72a-fb19-4a3a-8d4f-d22574868517', 'naam': 'tlc-fi-broker'},
            {'uuid': 'd9cf271a-91a3-413c-872a-db781d7177d1', 'naam': 'tlc-fi-broker'},
        ]
        qr = QueryResult(keys=['uuid', 'naam'], rows=rows, last_data_update=now, query_time_seconds=0.01)
        return qr

class DummySheetsWrapper:
    def read_data_from_sheet(self, spreadsheet_id, sheet_name, sheetrange=None, return_raw_results=False):
        # Provide empty mail receivers and empty historiek to keep behavior simple
        if return_raw_results:
            return {'values': [], 'range': f'{sheet_name}!A1:A1'}
        return []
    def get_sheets_in_spreadsheet(self, spreadsheet_id):
        # No Resultaat sheet present
        return {}
    def find_first_nonempty_row_from_starting_cell(self, spreadsheet_id, sheet_name, start_cell):
        return 2
    def read_celldata_from_sheet(self, spreadsheet_id, sheet_name, sheetrange=None):
        # minimal structure expected by DQReport summary code
        return {'startRow': 4, 'rowData': []}
    def insert_empty_rows(self, spreadsheet_id, sheet_name, start_cell, number_of_rows):
        return None
    def write_data_to_sheet(self, *args, **kwargs):
        # No-op: summary and historiek writes remain unfamiliar for Excel test
        return None

class DummyMailSender:
    def add_sheet_info(self, spreadsheet_id, mail_receivers_dict):
        pass
    def add_mail(self, *args, **kwargs):
        pass

# --- Patch factories and wrapper ---------------------------------------
orig_make_datasource = factories.make_datasource
orig_sheets_wrapper_get = None
try:
    import outputs.sheets_wrapper as sw
    orig_sheets_wrapper_get = sw.SingleSheetsWrapper.get_wrapper
except Exception:
    sw = None

factories.make_datasource = lambda name: DummyDatasource()
if sw:
    sw.SingleSheetsWrapper.get_wrapper = staticmethod(lambda: DummySheetsWrapper())

# --- Run the report ----------------------------------------------------
logging.basicConfig(level=logging.INFO)
report_class = Report0002()
report_class.init_report()
# run the report with dummy mail sender
sender = DummyMailSender()
try:
    report_class.run_report(sender=sender)
    # after run, locate file
    excel_name = report_class.report.excel_filename or report_class.report.title
    out_path = Path('RSA_OneDrive') / excel_name
    print('\nReport run completed; checking output file:')
    print(' ->', out_path)
    if out_path.exists():
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        print('sheets:', wb.sheetnames)
        ws = wb['Resultaat']
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            print(i+1, row)
            if i >= 20:
                break
    else:
        print('Output file not found')
finally:
    # restore
    factories.make_datasource = orig_make_datasource
    if sw and orig_sheets_wrapper_get:
        sw.SingleSheetsWrapper.get_wrapper = orig_sheets_wrapper_get

print('\nDone')

